"""XLSX/CSV parser using openpyxl and stdlib csv."""

import csv
import io
import logging
from pathlib import Path

from services.documents.parsers.pdf_parser import ParsedDocument, ParsedPage

logger = logging.getLogger(__name__)


def parse_xlsx(file_path: str | Path, file_name: str | None = None) -> ParsedDocument:
    """Extract text from XLSX files. Each sheet becomes a separate 'page'."""
    from openpyxl import load_workbook

    path = Path(file_path)
    fname = file_name or path.name
    wb = load_workbook(str(path), read_only=True, data_only=True)

    pages: list[ParsedPage] = []
    for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip().replace("|", "").strip():
                rows.append(row_text)
        if rows:
            # Prepend header with sheet name
            text = f"Sheet: {sheet_name}\n" + "\n".join(rows)
            pages.append(ParsedPage(page_number=sheet_idx, text=text))

    wb.close()
    logger.info(f"Parsed XLSX: {fname}, {len(pages)} sheets")

    return ParsedDocument(
        file_name=fname,
        pages=pages,
        total_pages=len(pages),
        metadata={"sheet_count": len(pages)},
    )


def parse_csv(file_path: str | Path, file_name: str | None = None) -> ParsedDocument:
    """Extract text from CSV files."""
    path = Path(file_path)
    fname = file_name or path.name

    with open(path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        rows = []
        for row in reader:
            row_text = " | ".join(row)
            if row_text.strip().replace("|", "").strip():
                rows.append(row_text)

    text = "\n".join(rows)
    pages = [ParsedPage(page_number=1, text=text)] if text.strip() else []

    logger.info(f"Parsed CSV: {fname}, {len(rows)} rows")

    return ParsedDocument(
        file_name=fname,
        pages=pages,
        total_pages=1,
        metadata={"row_count": len(rows)},
    )
